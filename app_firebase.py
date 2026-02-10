# =====================================================
# BACKEND FIREBASE - REPORTE SEMANAL/QUINCENAL
# Versión Profesional con Análisis de Demanda
# =====================================================
# Archivo: app.py

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import calendar
import os
import json
from dotenv import load_dotenv
from functools import wraps
import firebase_admin
from firebase_admin import credentials, firestore, storage, auth as firebase_auth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import io

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)
CORS(app)

# =====================================================
# INICIALIZACIÓN FIREBASE
# =====================================================

# Leer credenciales de Firebase desde variable de entorno
firebase_config = os.getenv('FIREBASE_CONFIG')

if firebase_config:
    try:
        config_dict = json.loads(firebase_config)
        cred = credentials.Certificate(config_dict)
        firebase_admin.initialize_app(cred, {
            'storageBucket': config_dict.get('storageBucket', '')
        })
        db = firestore.client()
        print("✅ Firebase inicializado exitosamente")
    except Exception as e:
        print(f"❌ Error al inicializar Firebase: {e}")
        db = None
else:
    print("⚠️  FIREBASE_CONFIG no configurado. Usa Cloud Firestore local o configura variables de entorno.")
    db = None

# =====================================================
# AUTENTICACIÓN Y PROTECCIÓN DE ENDPOINTS
# =====================================================

def verificar_token(f):
    """Decorador para verificar token de autenticación Firebase"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'error': 'Token requerido'}), 401
        
        try:
            # Remover "Bearer " del token si existe
            if token.startswith('Bearer '):
                token = token[7:]
            
            # Verificar token con Firebase Admin SDK
            decoded_token = firebase_auth.verify_id_token(token)
            request.user_id = decoded_token['uid']
            request.user_email = decoded_token['email']
            
            return f(*args, **kwargs)
        except Exception as e:
            return jsonify({'error': f'Token inválido: {str(e)}'}), 401
    
    return decorated_function

# =====================================================
# FUNCIONES DE FIREBASE FIRESTORE
# =====================================================

def inicializar_colecciones():
    """Inicializar colecciones y documentos predeterminados"""
    if not db:
        return
    
    try:
        # Verificar si medicamentos existen
        medicamentos_ref = db.collection('medicamentos')
        medicamentos_existentes = medicamentos_ref.stream()
        
        count = sum(1 for _ in medicamentos_existentes)
        
        if count == 0:
            print("📦 Creando medicamentos predeterminados...")
            medicamentos_orden = [
                "Acido valproico", "Amiodarona", "Atracurio", "Atropina",
                "Bicarbonato", "Clorfeniramida", "Clorudo de potasio", "Clorudo de sodio",
                "Dexametazona", "Diclofenaco", "Dicynone", "Diazepam", "Dimenhidrato",
                "Dipirona", "Dobutamina", "Efedrina", "Fentanil", "Flumazenil", "Fenitoina",
                "Fenobarbital", "Furosemida", "Gronisetron", "Gluconato de calcio",
                "Hidrocortizona", "Lidocaina", "Metilpredisona", "Metoclopramida",
                "Midazolan", "Morfina", "Norestimina", "Rosiverina", "Sulfato de magnecio"
            ]
            
            for idx, nombre in enumerate(medicamentos_orden):
                medicamentos_ref.add({
                    'nombre': nombre,
                    'stock_minimo': 10,
                    'orden': idx,
                    'fecha_creacion': datetime.now()
                })
            
            print("✅ Medicamentos creados")
    except Exception as e:
        print(f"❌ Error inicializando colecciones: {e}")

# =====================================================
# RUTAS DE MEDICAMENTOS
# =====================================================

@app.route('/api/medicamentos', methods=['GET'])
def get_medicamentos():
    """Obtener lista de medicamentos en orden"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        medicamentos = []
        docs = db.collection('medicamentos').order_by('orden').stream()
        
        for doc in docs:
            data = doc.to_dict()
            data['id'] = doc.id
            medicamentos.append(data)
        
        return jsonify(medicamentos)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/medicamentos', methods=['POST'])
def crear_medicamento():
    """Crear nuevo medicamento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        data = request.json
        nombre = data.get('nombre', '').strip()
        stock_minimo = data.get('stock_minimo', 10)
        
        if not nombre:
            return jsonify({'error': 'El nombre del medicamento es obligatorio'}), 400
        
        if stock_minimo < 1:
            return jsonify({'error': 'El stock mínimo debe ser mayor a 0'}), 400
        
        # Verificar si ya existe
        existing = db.collection('medicamentos').where('nombre', '==', nombre).stream()
        if sum(1 for _ in existing) > 0:
            return jsonify({'error': 'El medicamento ya existe'}), 409
        
        # Obtener siguiente orden
        docs = db.collection('medicamentos').order_by('orden', direction=firestore.Query.DESCENDING).limit(1).stream()
        max_orden = 0
        for doc in docs:
            max_orden = doc.to_dict().get('orden', 0)
        
        # Crear medicamento
        ref = db.collection('medicamentos').add({
            'nombre': nombre,
            'stock_minimo': stock_minimo,
            'orden': max_orden + 1,
            'fecha_creacion': datetime.now()
        })
        
        return jsonify({
            'id': ref[1].id,
            'nombre': nombre,
            'stock_minimo': stock_minimo,
            'mensaje': 'Medicamento agregado exitosamente'
        }), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/medicamentos/<id>', methods=['DELETE'])
def eliminar_medicamento_ruta(id):
    """Eliminar un medicamento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        # Verificar si tiene movimientos
        movimientos = db.collection('movimientos').where('medicamento_id', '==', id).stream()
        count = sum(1 for _ in movimientos)
        
        if count > 0:
            return jsonify({
                'error': f'No se puede eliminar. Hay {count} movimiento(s) asociado(s).'
            }), 400
        
        db.collection('medicamentos').document(id).delete()
        return jsonify({'mensaje': 'Medicamento eliminado exitosamente'}), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/medicamentos/<id>', methods=['PUT'])
def actualizar_medicamento_ruta(id):
    """Actualizar información de un medicamento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        data = request.json
        nombre = data.get('nombre', '').strip()
        stock_minimo = data.get('stock_minimo')
        
        if not nombre:
            return jsonify({'error': 'El nombre del medicamento es obligatorio'}), 400
        
        if stock_minimo is not None and stock_minimo < 1:
            return jsonify({'error': 'El stock mínimo debe ser mayor a 0'}), 400
        
        # Actualizar medicamento
        update_data = {'nombre': nombre}
        if stock_minimo is not None:
            update_data['stock_minimo'] = stock_minimo
        
        db.collection('medicamentos').document(id).update(update_data)
        return jsonify({'mensaje': 'Medicamento actualizado exitosamente'}), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# RUTAS DE MOVIMIENTOS
# =====================================================

@app.route('/api/movimientos', methods=['GET'])
def get_movimientos():
    """Obtener movimientos con filtros"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        limit = int(request.args.get('limit', 100))
        
        # Obtener movimientos sin filtros complejos (sin índices)
        query = db.collection('movimientos').order_by('fecha_registro', direction=firestore.Query.DESCENDING).limit(limit)
        
        movimientos = []
        for doc in query.stream():
            data = doc.to_dict()
            data['id'] = doc.id
            
            # Obtener nombre del medicamento
            med_id = data.get('medicamento_id')
            if med_id:
                med_doc = db.collection('medicamentos').document(med_id).get()
                if med_doc.exists:
                    data['medicamento_nombre'] = med_doc.to_dict().get('nombre', 'Desconocido')
                else:
                    data['medicamento_nombre'] = 'Desconocido'
            else:
                data['medicamento_nombre'] = 'Desconocido'
            
            movimientos.append(data)
        
        return jsonify(movimientos)
    
    except Exception as e:
        print(f"Error en get_movimientos: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos', methods=['POST'])
def crear_movimiento():
    """Registrar nuevo movimiento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        data = request.json
        
        tipo = data.get('tipo')
        fecha = data.get('fecha')
        medicamento_id = data.get('medicamento_id')
        cantidad = data.get('cantidad')
        
        if not all([tipo, fecha, medicamento_id, cantidad]):
            return jsonify({'error': 'Datos incompletos'}), 400
        
        if tipo not in ['INGRESO', 'SALIDA']:
            return jsonify({'error': 'Tipo inválido'}), 400
        
        if cantidad <= 0:
            return jsonify({'error': 'Cantidad debe ser mayor a 0'}), 400
        
        # Si es SALIDA, verificar stock
        if tipo == 'SALIDA':
            stock_actual = calcular_stock_medicamento(medicamento_id)
            if stock_actual < cantidad:
                return jsonify({
                    'error': 'Stock insuficiente',
                    'stock_actual': stock_actual,
                    'cantidad_solicitada': cantidad
                }), 400
        
        # Crear movimiento
        ref = db.collection('movimientos').add({
            'tipo': tipo,
            'fecha': fecha,
            'medicamento_id': medicamento_id,
            'cantidad': cantidad,
            'turno': data.get('turno'),
            'fecha_vencimiento': data.get('fecha_vencimiento'),
            'observaciones': data.get('observaciones'),
            'fecha_registro': datetime.now()
        })
        
        return jsonify({'id': ref[1].id, 'mensaje': 'Movimiento registrado'}), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos/<id>', methods=['DELETE'])
def eliminar_movimiento(id):
    """Eliminar un movimiento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        db.collection('movimientos').document(id).delete()
        return jsonify({'mensaje': 'Movimiento eliminado'})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/movimientos/<id>', methods=['PUT'])
def actualizar_movimiento(id):
    """Actualizar movimiento"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        data = request.json
        nueva_cantidad = data.get('cantidad')
        nueva_fecha = data.get('fecha')
        nuevo_turno = data.get('turno')
        
        if not nueva_cantidad or nueva_cantidad < 1:
            return jsonify({'error': 'Cantidad inválida'}), 400
        
        if not nueva_fecha:
            return jsonify({'error': 'Fecha requerida'}), 400
        
        # Obtener movimiento actual
        mov_doc = db.collection('movimientos').document(id).get()
        
        if not mov_doc.exists:
            return jsonify({'error': 'Movimiento no encontrado'}), 404
        
        movimiento = mov_doc.to_dict()
        
        # Actualizar movimiento
        update_data = {
            'cantidad': nueva_cantidad,
            'fecha': nueva_fecha
        }
        
        if nuevo_turno:
            update_data['turno'] = nuevo_turno
        
        db.collection('movimientos').document(id).update(update_data)
        return jsonify({'mensaje': 'Movimiento actualizado'}), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# INVENTARIO Y ANÁLISIS
# =====================================================

def calcular_stock_medicamento(medicamento_id):
    """Calcular stock actual de un medicamento"""
    if not db:
        return 0
    
    try:
        ingresos = 0
        salidas = 0
        
        # Obtener todos los movimientos del medicamento (sin filtro de tipo en where)
        for doc in db.collection('movimientos').where('medicamento_id', '==', medicamento_id).stream():
            mov = doc.to_dict()
            tipo = mov.get('tipo', '')
            cantidad = mov.get('cantidad', 0)
            
            if tipo == 'INGRESO':
                ingresos += cantidad
            elif tipo == 'SALIDA':
                salidas += cantidad
        
        return ingresos - salidas
    
    except Exception as e:
        print(f"Error calculando stock: {e}")
        return 0

@app.route('/api/inventario', methods=['GET'])
def get_inventario():
    """Obtener inventario completo"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        inventario = []
        
        for med_doc in db.collection('medicamentos').order_by('orden').stream():
            med = med_doc.to_dict()
            med_id = med_doc.id
            
            stock = calcular_stock_medicamento(med_id)
            stock_minimo = med.get('stock_minimo', 10)
            
            if stock <= 0:
                estado = 'AGOTADO'
            elif stock <= stock_minimo:
                estado = 'CRITICO'
            elif stock <= stock_minimo * 1.5:
                estado = 'BAJO'
            else:
                estado = 'OK'
            
            # Obtener último ingreso (obtener todos los movimientos y filtrar en Python)
            ultimo_ingreso = None
            for doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = doc.to_dict()
                if mov.get('tipo') == 'INGRESO':
                    if ultimo_ingreso is None or mov.get('fecha', '') > (ultimo_ingreso.get('fecha', '') or ''):
                        ultimo_ingreso = {
                            'fecha': mov.get('fecha'),
                            'lote': mov.get('observaciones')
                        }
            
            # Obtener egresos del mes (sin filtro de fecha en Firestore)
            hoy = datetime.now()
            primer_dia = datetime(hoy.year, hoy.month, 1).date().isoformat()
            ultimo_dia = (datetime(hoy.year, hoy.month, 1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
            ultimo_dia = ultimo_dia.date().isoformat()
            
            egresos_mes = 0
            # Obtener todos los movimientos del medicamento y filtrar en Python
            for doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = doc.to_dict()
                if mov.get('tipo') == 'SALIDA':
                    if mov.get('fecha', '') >= primer_dia and mov.get('fecha', '') <= ultimo_dia:
                        egresos_mes += mov.get('cantidad', 0)
            
            inventario.append({
                'id': med_id,
                'nombre': med.get('nombre'),
                'stock_actual': stock,
                'stock_minimo': stock_minimo,
                'estado': estado,
                'ultimo_ingreso': ultimo_ingreso,
                'egresos_mes': egresos_mes
            })
        
        return jsonify(inventario)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/estadisticas', methods=['GET'])
def get_estadisticas():
    """Obtener estadísticas generales"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        # Total medicamentos
        total_medicamentos = sum(1 for _ in db.collection('medicamentos').stream())
        
        # Alertas
        alertas = 0
        for med_doc in db.collection('medicamentos').stream():
            med = med_doc.to_dict()
            stock = calcular_stock_medicamento(med_doc.id)
            if stock <= med.get('stock_minimo', 10):
                alertas += 1
        
        # Movimientos hoy
        hoy = datetime.now().date().isoformat()
        movimientos_hoy = sum(1 for _ in db.collection('movimientos').where('fecha', '==', hoy).stream())
        
        return jsonify({
            'total_medicamentos': total_medicamentos,
            'alertas_stock_bajo': alertas,
            'movimientos_hoy': movimientos_hoy
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# REPORTES EXCEL
# =====================================================

def crear_reporte_excel(fecha_inicio, fecha_fin, titulo):
    """Crear reporte Excel con formato mejorado"""
    if not db:
        return None
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # CONFIGURAR ORIENTACIÓN HORIZONTAL
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        
        # ENCABEZADO
        ws.merge_cells('A1:Z1')
        ws['A1'] = 'REGISTRO DE MEDICAMENTOS DEL CARRO DE URGENCIAS'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = titulo
        ws['A2'].font = Font(bold=True, size=11)
        
        # ENCABEZADOS DE COLUMNAS
        row = 4
        
        ws.merge_cells(f'A{row}:A{row+1}')
        ws[f'A{row}'] = 'MEDICAMENTOS'
        
        ws.merge_cells(f'B{row}:B{row+1}')
        ws[f'B{row}'] = 'F. INGRESO'
        
        ws.merge_cells(f'C{row}:C{row+1}')
        ws[f'C{row}'] = 'F. VENC.'
        
        ws.merge_cells(f'D{row}:D{row+1}')
        ws[f'D{row}'] = 'STOCK INICIAL'
        
        # Calcular días
        dias = (fecha_fin - fecha_inicio).days + 1
        col = 5
        
        dias_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        
        for dia_offset in range(dias):
            fecha_dia = fecha_inicio + timedelta(days=dia_offset)
            dia_nombre = dias_semana[fecha_dia.weekday()]
            
            # Merge 3 columnas para cada día (M, T, N)
            ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}')
            ws[f'{get_column_letter(col)}{row}'] = f'{dia_nombre}\n{fecha_dia.strftime("%d/%m")}'
            ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Turnos M, T, N
            ws[f'{get_column_letter(col)}{row+1}'] = 'M'
            ws[f'{get_column_letter(col+1)}{row+1}'] = 'T'
            ws[f'{get_column_letter(col+2)}{row+1}'] = 'N'
            
            col += 3
        
        # Columnas finales: Demanda Total y Stock Actual
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col)}{row+1}')
        ws[f'{get_column_letter(col)}{row}'] = 'DEMANDA TOTAL'
        col_demanda = col
        col += 1
        
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col)}{row+1}')
        ws[f'{get_column_letter(col)}{row}'] = 'STOCK ACTUAL'
        col_stock = col
        
        # Aplicar estilo a encabezados
        for r in range(row, row+2):
            for c in range(1, col_stock+2):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        ws.row_dimensions[row].height = 35
        ws.row_dimensions[row+1].height = 20
        
        # DATOS DE MEDICAMENTOS
        data_row = row + 2
        
        for med_doc in db.collection('medicamentos').order_by('orden').stream():
            med = med_doc.to_dict()
            med_id = med_doc.id
            
            ws[f'A{data_row}'] = med.get('nombre')
            ws[f'A{data_row}'].font = Font(size=9)
            
            # Último ingreso - query simple con filtrado en Python
            ultimo_ingreso = None
            for mov_doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = mov_doc.to_dict()
                if mov.get('tipo') == 'INGRESO':
                    if ultimo_ingreso is None or mov.get('fecha', '') > (ultimo_ingreso.get('fecha', '') or ''):
                        ultimo_ingreso = mov
            
            if ultimo_ingreso:
                ws[f'B{data_row}'] = ultimo_ingreso.get('fecha')
                if ultimo_ingreso.get('fecha_vencimiento'):
                    ws[f'C{data_row}'] = ultimo_ingreso.get('fecha_vencimiento')
            
            # Stock inicial
            stock_inicial = calcular_stock_medicamento(med_id)
            ws[f'D{data_row}'] = stock_inicial
            ws[f'D{data_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Salidas por día y turno - query simple con filtrado en Python
            col_idx = 5
            
            # Obtener todos los movimientos de salida para este medicamento
            movimientos_salida = []
            for mov_doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = mov_doc.to_dict()
                if mov.get('tipo') == 'SALIDA':
                    movimientos_salida.append(mov)
            
            for dia_offset in range(dias):
                fecha_dia = fecha_inicio + timedelta(days=dia_offset)
                
                for turno in ['M', 'T', 'N']:
                    total = 0
                    for mov in movimientos_salida:
                        if mov.get('fecha') == str(fecha_dia) and mov.get('turno') == turno:
                            total += mov.get('cantidad', 0)
                    
                    if total > 0:
                        ws.cell(row=data_row, column=col_idx).value = total
                        ws.cell(row=data_row, column=col_idx).font = Font(size=9)
                    
                    col_idx += 1
            
            # Demanda Total - sumar del período (egresos solo)
            demanda = 0
            for mov in movimientos_salida:
                mov_fecha = mov.get('fecha', '')
                if mov_fecha >= str(fecha_inicio) and mov_fecha <= str(fecha_fin):
                    demanda += mov.get('cantidad', 0)
            
            ws.cell(row=data_row, column=col_demanda).value = demanda
            ws.cell(row=data_row, column=col_demanda).font = Font(size=9, bold=True)
            
            # Stock Actual
            stock_actual = calcular_stock_medicamento(med_id)
            ws.cell(row=data_row, column=col_stock).value = stock_actual
            ws.cell(row=data_row, column=col_stock).font = Font(size=9, bold=True)
            
            # Bordes
            for c in range(1, col_stock+1):
                cell = ws.cell(row=data_row, column=c)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if c > 5:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            data_row += 1
        
        # PIE
        ws[f'A{data_row+1}'] = 'Observaciones: Los medicamentos deben ser seleccionados según necesidades del servicio'
        ws[f'A{data_row+1}'].font = Font(size=8)
        
        ws[f'B{data_row+3}'] = 'FIRMA RESPONSABLE:'
        ws[f'B{data_row+3}'].font = Font(bold=True, size=9)
        ws[f'B{data_row+4}'] = '_____________________'
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        for col in range(5, col_stock+1):
            ws.column_dimensions[get_column_letter(col)].width = 5
        
        # Configurar márgenes para landscape
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        
        return wb
    
    except Exception as e:
        print(f"Error creando reporte: {e}")
        return None

@app.route('/api/reportes/semanal-excel', methods=['GET'])
def generar_reporte_semanal_excel():
    """Generar reporte semanal en Excel"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        
        if not fecha_inicio:
            return jsonify({'error': 'Fecha de inicio requerida'}), 400
        
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = fecha_inicio + timedelta(days=6)
        
        titulo = f'Semana del {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
        
        wb = crear_reporte_excel(fecha_inicio, fecha_fin, titulo)
        
        if not wb:
            return jsonify({'error': 'Error creando reporte'}), 500
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        nombre_archivo = f'Reporte_Semanal_{fecha_inicio.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/quincenal-excel', methods=['GET'])
def generar_reporte_quincenal_excel():
    """Generar reporte quincenal en Excel"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        
        if not fecha_inicio:
            return jsonify({'error': 'Fecha de inicio requerida'}), 400
        
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = fecha_inicio + timedelta(days=14)
        
        titulo = f'Quincena del {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
        
        wb = crear_reporte_excel(fecha_inicio, fecha_fin, titulo)
        
        if not wb:
            return jsonify({'error': 'Error creando reporte'}), 500
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        nombre_archivo = f'Reporte_Quincenal_{fecha_inicio.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# ANÁLISIS Y REPORTES
# =====================================================

@app.route('/api/analisis/demanda', methods=['GET'])
def analisis_demanda():
    """Análisis de demanda de medicamentos"""
    try:
        if not db:
            return jsonify({'error': 'Base de datos no disponible'}), 500
        
        dias = int(request.args.get('dias', 30))
        
        # Calcular fecha desde
        fecha_desde = (datetime.now() - timedelta(days=dias)).date().isoformat()
        
        # Analizar demanda
        demanda = {}
        
        for mov_doc in db.collection('movimientos').where('tipo', '==', 'SALIDA').stream():
            mov = mov_doc.to_dict()
            if mov.get('fecha', '') >= fecha_desde:
                med_id = mov.get('medicamento_id')
                if med_id not in demanda:
                    demanda[med_id] = {
                        'cantidad': 0,
                        'frecuencia': 0,
                        'nombre': ''
                    }
                demanda[med_id]['cantidad'] += mov.get('cantidad', 0)
                demanda[med_id]['frecuencia'] += 1
        
        # Obtener nombres y ordenar
        resultado = []
        for med_id, data in demanda.items():
            med_doc = db.collection('medicamentos').document(med_id).get()
            if med_doc.exists:
                nombre = med_doc.to_dict().get('nombre', 'Desconocido')
                resultado.append({
                    'nombre': nombre,
                    'total_dispensado': data['cantidad'],
                    'frecuencia': data['frecuencia'],
                    'promedio_diario': round(data['cantidad'] / dias, 2)
                })
        
        # Ordenar por cantidad descendente
        resultado.sort(key=lambda x: x['total_dispensado'], reverse=True)
        
        return jsonify(resultado[:10])  # Top 10
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/semanal-pdf', methods=['GET'])
def generar_reporte_semanal_pdf():
    """Generar reporte semanal en PDF"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        
        if not fecha_inicio:
            return jsonify({'error': 'Fecha de inicio requerida'}), 400
        
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = fecha_inicio_dt + timedelta(days=6)
        
        # Crear PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), topMargin=15, bottomMargin=15)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Contenido
        elements = []
        
        # Título
        title = Paragraph(f"<b>Reporte Semanal: {fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Crear tabla
        data = [['Medicamento', 'Stock Inicial', 'Dem. Lun', 'Dem. Mar', 'Dem. Mié', 'Dem. Jue', 'Dem. Vie', 'Dem. Sáb', 'Dem. Dom', 'Total Demanda', 'Stock Final']]
        
        for med_doc in db.collection('medicamentos').order_by('orden').limit(20).stream():
            med = med_doc.to_dict()
            med_id = med_doc.id
            
            stock_inicial = calcular_stock_medicamento(med_id)
            row = [med.get('nombre', 'N/A'), stock_inicial]
            
            # Obtener todos los movimientos de salida para este medicamento
            movimientos_salida = []
            for mov_doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = mov_doc.to_dict()
                if mov.get('tipo') == 'SALIDA':
                    movimientos_salida.append(mov)
            
            total_demanda = 0
            
            for i in range(7):
                fecha_dia = fecha_inicio_dt + timedelta(days=i)
                demanda_dia = 0
                for mov in movimientos_salida:
                    if mov.get('fecha') == str(fecha_dia):
                        demanda_dia += mov.get('cantidad', 0)
                
                row.append(demanda_dia)
                total_demanda += demanda_dia
            
            row.append(total_demanda)
            stock_final = stock_inicial - total_demanda
            row.append(stock_final)
            
            data.append(row)
        
        # Crear tabla
        table = Table(data, colWidths=[2*inch, 0.6*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
        ]))
        
        elements.append(table)
        
        # Generar PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        
        nombre_archivo = f'Reporte_Semanal_{fecha_inicio_dt.strftime("%Y%m%d")}.pdf'
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre_archivo
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/quincenal-pdf', methods=['GET'])
def generar_reporte_quincenal_pdf():
    """Generar reporte quincenal en PDF"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        
        if not fecha_inicio:
            return jsonify({'error': 'Fecha de inicio requerida'}), 400
        
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = fecha_inicio_dt + timedelta(days=14)
        
        # Crear PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), topMargin=15, bottomMargin=15)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Contenido
        elements = []
        
        # Título
        title = Paragraph(f"<b>Reporte Quincenal: {fecha_inicio_dt.strftime('%d/%m/%Y')} al {fecha_fin_dt.strftime('%d/%m/%Y')}</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Crear tabla
        data = [['Medicamento', 'Stock Inicial', 'Demanda Total', 'Stock Final']]
        
        for med_doc in db.collection('medicamentos').order_by('orden').stream():
            med = med_doc.to_dict()
            med_id = med_doc.id
            
            stock_inicial = calcular_stock_medicamento(med_id)
            
            total_demanda = 0
            for mov_doc in db.collection('movimientos').where('medicamento_id', '==', med_id).stream():
                mov = mov_doc.to_dict()
                if mov.get('tipo') == 'SALIDA' and mov.get('fecha', '') >= str(fecha_inicio_dt) and mov.get('fecha', '') <= str(fecha_fin_dt):
                    total_demanda += mov.get('cantidad', 0)
            
            stock_final = stock_inicial - total_demanda
            
            data.append([med.get('nombre', 'N/A'), stock_inicial, total_demanda, stock_final])
        
        # Crear tabla
        table = Table(data, colWidths=[3*inch, 1*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        
        # Generar PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        
        nombre_archivo = f'Reporte_Quincenal_{fecha_inicio_dt.strftime("%Y%m%d")}.pdf'
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nombre_archivo
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# HEALTH CHECK
# =====================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Verificar estado de la aplicación"""
    return jsonify({
        'status': 'ok',
        'firebase': 'connected' if db else 'disconnected'
    })

@app.route('/api/test/data', methods=['POST'])
def crear_datos_prueba():
    """Crear datos de prueba (solo para desarrollo)"""
    if not db:
        return jsonify({'error': 'BD no disponible'}), 500
    
    try:
        # Crear medicamentos
        medicamentos_orden = [
            "Acido valproico", "Amiodarona", "Atracurio", "Atropina",
            "Bicarbonato", "Clorfeniramida", "Cloruro de potasio", "Cloruro de sodio",
            "Dexametazona", "Diclofenaco", "Dicynone", "Diazepam", "Dimenhidrato",
            "Dipirona", "Dobutamina", "Efedrina", "Fentanil", "Flumazenil", "Fenitoina",
            "Fenobarbital", "Furosemida", "Gronisetron", "Gluconato de calcio",
            "Hidrocortizona", "Lidocaina", "Metilpredisona", "Metoclopramida",
            "Midazolan", "Morfina", "Norestimina", "Rosiverina", "Sulfato de magnecio"
        ]
        
        med_ids = []
        for idx, nombre in enumerate(medicamentos_orden):
            ref = db.collection('medicamentos').add({
                'nombre': nombre,
                'stock_minimo': 10,
                'orden': idx,
                'fecha_creacion': datetime.now()
            })
            med_ids.append(ref[1].id)
        
        # Crear algunos movimientos de prueba
        hoy = datetime.now().date().isoformat()
        for i in range(min(5, len(med_ids))):
            db.collection('movimientos').add({
                'tipo': 'INGRESO',
                'fecha': hoy,
                'medicamento_id': med_ids[i],
                'cantidad': 50 + (i * 10),
                'turno': 'M',
                'fecha_vencimiento': None,
                'observaciones': f'Ingreso inicial {i+1}',
                'fecha_registro': datetime.now()
            })
        
        return jsonify({
            'mensaje': 'Datos de prueba creados',
            'medicamentos': len(med_ids),
            'movimientos': min(5, len(med_ids))
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =====================================================
# INICIALIZACIÓN
# =====================================================

if __name__ == '__main__':
    if db:
        inicializar_colecciones()
    
    print("""
    ╔════════════════════════════════════════════════╗
    ║  INVENTARIO MEDICAMENTOS - FIREBASE            ║
    ║  Versión Profesional con Firestore             ║
    ╠════════════════════════════════════════════════╣
    ║  Acceso Local:  http://localhost:5000          ║
    ║                                                ║
    ║  Características:                               ║
    ║  • Firestore Database                          ║
    ║  • Reporte Semanal y Quincenal                 ║
    ║  • Exportación Excel                           ║
    ║  • Análisis de demanda                         ║
    ║  • Compatible con Firebase Hosting             ║
    ╚════════════════════════════════════════════════╝
    """)
    
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
